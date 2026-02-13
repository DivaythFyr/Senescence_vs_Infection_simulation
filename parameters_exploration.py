"""
Comprehensive Parameter Exploration Script with Excel Export

Systematically tests ALL parameters in parameters.py to find configurations
where the infection wins.

Strategy:
1. Sequential Greedy Optimization:
   - Starts with original parameters.
   - Iterates through the parameter list (prioritizing user requests).
   - For each parameter, tests a wide range of values (0.1x to 10x).
   - If a value triggers an "INFECTION WIN", it is LOCKED IN as the new baseline
     for subsequent tests.
2. Supports both Integer and Float parameters automatically.
3. Skips structural constants (GRID_SIZE, MAX_TIME).
4. Exports all results to Excel file for easy analysis.
"""

import numpy as np
import parameters
import inspect
import time
import sys
from simulation_core import MatrixSimulation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# --- CONFIGURATION ---

# Structural constants to SKIP (do not optimize these)
SKIP_PARAMS = {
    'GRID_SIZE', 
    'MAX_SIMULATION_TIMEPOINTS'
}

# Parameters to prioritize (tested first)
PRIORITY_ORDER = [
    'PEAK_VIRUS_PRODUCTION_PART_OF_LIFESPAN',
    'INFECTED_LIFE_LIMIT',
    'INFECTIVITY',
    'VIRUS_PRODUCTION_START_RATIO',
    'BASE_VIRUS_RELEASE'
]

# Test multipliers to generate ranges around original values
MULTIPLIERS = [0.1, 0.5, 0.8, 1.2, 1.5, 2.0, 5.0, 10.0]

# --- HELPER FUNCTIONS ---

def get_adjustable_parameters():
    """Extracts all parameter names and values from parameters.py, excluding skips."""
    members = inspect.getmembers(parameters)
    params = {}
    for name, value in members:
        if name.isupper() and not name.startswith('_') and name not in SKIP_PARAMS:
            params[name] = value
    return params

def sort_parameters(param_names):
    """Sorts parameters: Priority ones first, then alphabetical."""
    sorted_list = []
    # Add priority items if they exist
    for p in PRIORITY_ORDER:
        if p in param_names:
            sorted_list.append(p)

    # Add the rest
    remaining = sorted([p for p in param_names if p not in sorted_list])
    return sorted_list + remaining

def generate_test_values(original_val):
    """Generates a list of test values based on multipliers."""
    is_int = isinstance(original_val, int)
    values = []

    if original_val == 0:
        # Special case for 0: test small positive numbers
        if is_int:
            values = [1, 2, 5, 10]
        else:
            values = [0.01, 0.05, 0.1, 0.5, 1.0]
    else:
        for m in MULTIPLIERS:
            new_val = original_val * m
            if is_int:
                new_val = int(round(new_val))
            values.append(new_val)

    unique_values = sorted(list(set([v for v in values if v >= 0])))

    # Remove the original value if present to avoid redundant testing
    if original_val in unique_values:
        unique_values.remove(original_val)

    return unique_values

def run_simulation_with_overrides(overrides):
    """
    Runs a simulation with specific parameter overrides.
    Returns the outcome result dict.
    """
    # Apply overrides
    original_values = {}
    for k, v in overrides.items():
        if hasattr(parameters, k):
            original_values[k] = getattr(parameters, k)
            setattr(parameters, k, v)

    # Run Simulation
    sim = MatrixSimulation(parameters.GRID_SIZE)
    outcome = "UNCLEAR"

    # Run
    for _ in range(parameters.MAX_SIMULATION_TIMEPOINTS):
        if not sim.run_time_step():
            break

    # Restore parameters immediately
    for k, v in original_values.items():
        setattr(parameters, k, v)

    # Analyze Result
    counts = sim.grid.get_cell_counts()
    infected = counts.get('infected', 0)
    empty = counts.get('empty', 0)
    total_cells = parameters.GRID_SIZE ** 2

    # Definition of Infection Win:
    if infected > 0 or empty > 0.8 * total_cells:
        outcome = "INFECTION_WIN"
    elif (counts.get('susceptible', 0) + counts.get('resistant', 0) + counts.get('senescent', 0)) > 0.5 * total_cells:
        outcome = "HEALTHY_WIN"

    return {
        'outcome': outcome,
        'infected': infected,
        'empty': empty,
        'susceptible': counts.get('susceptible', 0),
        'resistant': counts.get('resistant', 0),
        'senescent': counts.get('senescent', 0),
        'duration': time.time() - time.time(),
        'final_time': sim.time
    }

def export_to_excel(results_list, filename='parameter_exploration_results.xlsx'):
    """
    Exports exploration results to an Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    infection_win_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    healthy_win_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    unclear_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ['Test #', 'Parameter', 'Test Value', 'Original Value', 'Outcome', 
               'Infected Cells', 'Empty Cells', 'Susceptible', 'Resistant', 'Senescent', 
               'Duration (s)', 'Final Timepoint']

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Add data rows
    for row_idx, result in enumerate(results_list, 2):
        cells_data = [
            result.get('test_num'),
            result.get('parameter'),
            result.get('test_value'),
            result.get('original_value'),
            result.get('outcome'),
            result.get('infected'),
            result.get('empty'),
            result.get('susceptible'),
            result.get('resistant'),
            result.get('senescent'),
            round(result.get('duration', 0), 2),
            result.get('final_time')
        ]

        for col_idx, value in enumerate(cells_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = center_align

            # Color code by outcome
            if result.get('outcome') == 'INFECTION_WIN':
                cell.fill = infection_win_fill
                cell.font = Font(bold=True)
            elif result.get('outcome') == 'HEALTHY_WIN':
                cell.fill = healthy_win_fill
            elif result.get('outcome') == 'UNCLEAR':
                cell.fill = unclear_fill

    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 18

    # Add Summary Sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws.column_dimensions['A'].width = 30
    summary_ws.column_dimensions['B'].width = 20

    # Count outcomes
    infection_wins = sum(1 for r in results_list if r['outcome'] == 'INFECTION_WIN')
    healthy_wins = sum(1 for r in results_list if r['outcome'] == 'HEALTHY_WIN')
    unclear = sum(1 for r in results_list if r['outcome'] == 'UNCLEAR')

    summary_data = [
        ('Exploration Summary', ''),
        ('Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('Total Tests', len(results_list)),
        ('Infection Win', infection_wins),
        ('Healthy Win', healthy_wins),
        ('Unclear', unclear),
        ('', ''),
        ('Infection Win Details', ''),
    ]

    row = 1
    for label, value in summary_data:
        cell_a = summary_ws.cell(row=row, column=1)
        cell_a.value = label
        cell_a.font = Font(bold=True)

        cell_b = summary_ws.cell(row=row, column=2)
        cell_b.value = value

        row += 1

    # Add infection win details
    infection_win_results = [r for r in results_list if r['outcome'] == 'INFECTION_WIN']
    if infection_win_results:
        row += 1
        for result in infection_win_results:
            summary_ws.cell(row=row, column=1).value = f"{result['parameter']} = {result['test_value']}"
            summary_ws.cell(row=row, column=2).value = f"(Original: {result['original_value']})"
            row += 1

    # Save
    wb.save(filename)
    return filename

# --- MAIN EXECUTION ---

def main():
    print("="*80)
    print("COMPREHENSIVE PARAMETER EXPLORATION WITH EXCEL EXPORT")
    print("Goal: Find ANY parameter set where Infection Wins")
    print("="*80)

    # 1. Initialize Baseline
    current_best_params = get_adjustable_parameters()
    param_names = sort_parameters(list(current_best_params.keys()))

    print(f"\nLoaded {len(param_names)} parameters to test.")
    print("Baseline: Original parameters (Presumed Healthy Win)")

    # Storage for all results
    all_results = []
    test_counter = 0

    # Verify baseline
    print("\nVerifying baseline...")
    base_res = run_simulation_with_overrides({})
    print(f"Baseline Outcome: {base_res['outcome']}")

    if base_res['outcome'] == "INFECTION_WIN":
        print("Original parameters already lead to Infection Win.")

    # 2. Iterative Optimization
    found_any_improvement = False

    for param in param_names:
        original_val = current_best_params[param]
        test_values = generate_test_values(original_val)

        print(f"\n[{param}] Testing {len(test_values)} variations around {original_val}...")

        found_better_here = False

        for val in test_values:
            test_counter += 1

            # Construct full override set
            overrides = current_best_params.copy()
            overrides[param] = val

            res = run_simulation_with_overrides(overrides)

            # Store result
            result_record = {
                'test_num': test_counter,
                'parameter': param,
                'test_value': val,
                'original_value': original_val,
                'outcome': res['outcome'],
                'infected': res['infected'],
                'empty': res['empty'],
                'susceptible': res['susceptible'],
                'resistant': res['resistant'],
                'senescent': res['senescent'],
                'duration': res['duration'],
                'final_time': res['final_time']
            }
            all_results.append(result_record)

            # Status dot
            sys.stdout.write(".")
            sys.stdout.flush()

            if res['outcome'] == "INFECTION_WIN":
                print(f"\n  !!! DISCOVERY #{test_counter}: {param} = {val} leads to INFECTION_WIN !!!")

                # Update the baseline immediately (Greedy approach)
                current_best_params[param] = val
                found_any_improvement = True
                found_better_here = True
                break

        if not found_better_here:
            print(f" [No improvement]")

    # 3. Export to Excel
    print(f"\n\nExporting {len(all_results)} test results to Excel...")
    filename = export_to_excel(all_results)
    print(f"✓ Results saved to: {filename}")

    # 4. Final Report
    print("\n" + "="*80)
    print("EXPLORATION COMPLETE")
    print("="*80)
    print(f"Total tests run: {test_counter}")
    print(f"Infection wins found: {sum(1 for r in all_results if r['outcome'] == 'INFECTION_WIN')}")
    print(f"Healthy wins: {sum(1 for r in all_results if r['outcome'] == 'HEALTHY_WIN')}")
    print(f"Unclear: {sum(1 for r in all_results if r['outcome'] == 'UNCLEAR')}")

    if found_any_improvement:
        print("\nSUCCESS! Found parameter configuration(s) where Infection Wins.")
        print("Modified Parameters:")

        defaults = get_adjustable_parameters()
        for k, v in current_best_params.items():
            if v != defaults[k]:
                print(f"  {k} = {v}  # Original: {defaults[k]}")
    else:
        print("\nNo single parameter change flipped the outcome to Infection Win.")

    print(f"\n✓ Check '{filename}' for detailed results!")

if __name__ == "__main__":
    main()
